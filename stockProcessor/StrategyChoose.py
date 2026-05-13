import pandas as pd
import tushare as ts
import os
from copy import deepcopy
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from module.basic import basic_api
from module.config import tushare_config
from stockProcessor import eps_download as eps_downloader
from stockProcessor.download.constants import data_path, score_result_path
import strategy_choose_config

# ======================
# 参数区
# ======================
def load_strategy_runtime_config():
    preset_name = strategy_choose_config.ACTIVE_STRATEGY
    preset = strategy_choose_config.STRATEGY_PRESETS.get(preset_name)
    if preset is None:
        raise ValueError(f"未找到策略配置: {preset_name}")

    config = deepcopy(strategy_choose_config.DEFAULT_CONFIG)
    preset_conditions = preset.get("conditions", {})
    config["conditions"].update(preset_conditions)
    for key, value in preset.items():
        if key == "conditions":
            continue
        config[key] = value

    config["strategy_name"] = preset_name
    return config


RUNTIME_CONFIG = load_strategy_runtime_config()
END_DATE = None  # None 表示自动取最近一个已结束的交易日
MARKET_CLOSE_HOUR = int(RUNTIME_CONFIG.get("market_close_hour", 15))
DATA_LOOKBACK_TRADE_DAYS = RUNTIME_CONFIG["data_lookback_trade_days"]
SIGNAL_DAYS = RUNTIME_CONFIG["signal_days"]
MAX_FORWARD_DAYS = RUNTIME_CONFIG["max_forward_days"]
VOL_MA_DAYS = RUNTIME_CONFIG["vol_ma_days"]
RSI_PERIOD = RUNTIME_CONFIG["rsi_period"]
DAILY_CACHE_CSV = data_path("strategy_daily_cache.csv")
BASIC_CACHE_CSV = data_path("strategy_basic_cache.csv")
SHAREHOLDER_CACHE_CSV = data_path("strategy_shareholder_cache.csv")
MONEYFLOW_CACHE_CSV = data_path("strategy_moneyflow_cache.csv")
MIN_DV_TTM = RUNTIME_CONFIG.get("min_dv_ttm")
MAX_DV_TTM = RUNTIME_CONFIG.get("max_dv_ttm")
RESULT_XLSX = score_result_path("strategy_choose_result.xlsx")
CONDITION_FLAGS = RUNTIME_CONFIG["conditions"]
RSI_MAX = RUNTIME_CONFIG["rsi_max"]
NEAR_MA_THRESHOLD = RUNTIME_CONFIG["near_ma_threshold"]
RECENT_HIGH_LOOKBACK_DAYS = RUNTIME_CONFIG["recent_high_lookback_days"]
def normalize_optional_number(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return float(value)


MIN_EPS = normalize_optional_number(RUNTIME_CONFIG["min_eps"])
EPS_FILTER_ENABLED = bool(RUNTIME_CONFIG.get("enable_eps_filter", True)) and MIN_EPS is not None
MIN_TOTAL_MV = normalize_optional_number(RUNTIME_CONFIG.get("min_total_mv"))
TOTAL_MV_FILTER_ENABLED = bool(RUNTIME_CONFIG.get("enable_total_mv_filter", True)) and MIN_TOTAL_MV is not None
MIN_VOLUME_RATIO = RUNTIME_CONFIG["min_volume_ratio"]
MIN_EXTERNAL_INTERNAL_RATIO = RUNTIME_CONFIG["min_external_internal_ratio"]
MIN_TURNOVER_RATE = RUNTIME_CONFIG["min_turnover_rate"]
MAX_TURNOVER_RATE = RUNTIME_CONFIG["max_turnover_rate"]
MIN_PE = 0.0
MAX_PE = normalize_optional_number(RUNTIME_CONFIG["max_pe"])
PE_FILTER_ENABLED = bool(CONDITION_FLAGS.get("pe_reasonable", False)) and MAX_PE is not None
INDUSTRY_PE_MIN_SAMPLE_COUNT = int(RUNTIME_CONFIG.get("industry_pe_min_sample_count", 5))
INDUSTRY_PE_MAX_PERCENTILE = float(RUNTIME_CONFIG.get("industry_pe_max_percentile", 0.35))
INDUSTRY_PE_MAX_RATIO_TO_MEDIAN = float(RUNTIME_CONFIG.get("industry_pe_max_ratio_to_median", 0.8))

CONDITION_NAMES = {
    "trend_above_ma20": "股价在20日线之上",
    "bullish_ma_alignment": "5日线 > 10日线 > 20日线",
    "volume_rule": "上涨放量或回调缩量",
    "position_rule": "回踩10/20日线缩量，或突破前高放量",
    "macd_golden_cross": "MACD金叉",
    "rsi_not_overheated": "RSI不过热",
    "volume_ratio_high": "量比达标",
    "external_internal_ratio_high": "外盘/内盘达标",
    "turnover_rate_range": "换手率在设定区间",
    "pe_reasonable": "市盈率合理",
    "industry_relative_valuation_low": "相对行业估值偏低",
    "social_security_holder": "股东成分包含全国社保基金",
    "prev_year_high_dividend": "TTM股息率在区间内",
    "main_money_inflow_2days": "主力资金连续流入2天",
}

# ======================
# 工具函数
# ======================
def pro_api():
    return ts.pro_api(tushare_config.TuShareConst.TOKEN)


def fetch_with_retry(fetch_func, label):
    try:
        return fetch_func()
    except Exception as e:
        raise RuntimeError(f"{label} 下载失败，已停止本次任务，避免使用不完整数据：{e}") from e


def get_trade_dates(end_date, count=80):
    pro = pro_api()
    calendar_lookback_days = max(180, count * 2 + 30)
    cal = pro.trade_cal(
        exchange="SSE",
        start_date=(datetime.strptime(end_date, "%Y%m%d") - timedelta(days=calendar_lookback_days)).strftime("%Y%m%d"),
        end_date=end_date,
        is_open="1"
    )
    cal = cal.sort_values("cal_date")
    return cal["cal_date"].tolist()[-count:]


def resolve_trade_date_with_required_data(target_date, require_moneyflow=False, lookback_count=20):
    """
    从 target_date 往前寻找“所需数据都已就绪”的最近交易日。
    daily 是基础要求；当策略依赖资金流条件时，moneyflow 也必须可用。
    """
    candidate_dates = get_trade_dates(target_date, count=max(2, lookback_count))
    if not candidate_dates:
        raise ValueError(f"未获取到 {target_date} 之前的可用交易日，请检查交易日历接口")

    latest_candidate = candidate_dates[-1]
    pro = pro_api()
    required_data_desc = "daily"
    if require_moneyflow:
        required_data_desc = "daily + moneyflow"

    for trade_date in reversed(candidate_dates):
        daily_df = pro.daily(trade_date=trade_date)
        if daily_df.empty:
            continue
        if require_moneyflow:
            moneyflow_df = pro.moneyflow(trade_date=trade_date)
            if moneyflow_df.empty:
                continue
        if trade_date != latest_candidate:
            print(f"目标交易日 {latest_candidate} 的 {required_data_desc} 数据未就绪，回退到 {trade_date}")
        return trade_date

    raise ValueError(f"最近交易日均未获取到可用的 {required_data_desc} 数据，请检查 TuShare 接口")


def get_latest_completed_trade_date(now=None, require_moneyflow=False):
    if now is None:
        now = datetime.now()

    today = now.strftime("%Y%m%d")
    pro = pro_api()
    cal = pro.trade_cal(
        exchange="SSE",
        start_date=(now - timedelta(days=30)).strftime("%Y%m%d"),
        end_date=today,
        is_open="1"
    )
    cal = cal.sort_values("cal_date")
    trade_dates = cal["cal_date"].tolist()
    if not trade_dates:
        raise ValueError("未获取到可用交易日，请检查交易日历接口")

    if trade_dates[-1] != today or now.hour >= MARKET_CLOSE_HOUR:
        latest_candidate = trade_dates[-1]
    else:
        if len(trade_dates) < 2:
            raise ValueError("当前交易日未结束，且没有可用的上一交易日")
        latest_candidate = trade_dates[-2]

    return resolve_trade_date_with_required_data(
        latest_candidate,
        require_moneyflow=require_moneyflow,
    )


def calc_rsi(close, period=6):
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)

    avg_gain = gain.rolling(period).mean()
    avg_loss = loss.rolling(period).mean()

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi


def calc_macd(close, fast=12, slow=26, signal=9):
    ema_fast = close.ewm(span=fast, adjust=False).mean()
    ema_slow = close.ewm(span=slow, adjust=False).mean()
    dif = ema_fast - ema_slow
    dea = dif.ewm(span=signal, adjust=False).mean()
    macd = (dif - dea) * 2
    return dif, dea, macd


def get_moneyflow(ts_code, start_date, end_date):
    try:
        pro = pro_api()
        mf = pro.moneyflow(
            ts_code=ts_code,
            start_date=start_date,
            end_date=end_date
        )
        if mf.empty:
            return None

        mf = mf.sort_values("trade_date")
        return mf
    except Exception:
        return None


def is_main_money_inflow_2days(ts_code, trade_date, trade_dates):
    """
    判断主力资金连续流入2天。
    这里用 buy_lg_amount + buy_elg_amount - sell_lg_amount - sell_elg_amount 作为主力净流入。
    """
    idx = trade_dates.index(trade_date)
    if idx < 1:
        return False

    start_date = trade_dates[idx - 1]
    end_date = trade_date

    mf = get_moneyflow(ts_code, start_date, end_date)
    if mf is None or len(mf) < 2:
        return False

    mf["main_net"] = (
        mf["buy_lg_amount"].fillna(0)
        + mf["buy_elg_amount"].fillna(0)
        - mf["sell_lg_amount"].fillna(0)
        - mf["sell_elg_amount"].fillna(0)
    )

    last2 = mf.tail(2)
    return (last2["main_net"] > 0).all()


def add_daily_indicators(df):
    if df.empty:
        return df

    df = df.sort_values("trade_date").reset_index(drop=True)

    df["ma5"] = df["close"].rolling(5).mean()
    df["ma10"] = df["close"].rolling(10).mean()
    df["ma20"] = df["close"].rolling(20).mean()
    df["vol_ma5"] = df["vol"].rolling(VOL_MA_DAYS).mean()

    df["rsi"] = calc_rsi(df["close"], RSI_PERIOD)
    df["dif"], df["dea"], df["macd"] = calc_macd(df["close"])

    df["macd_gold"] = (df["dif"] > df["dea"]) & (df["dif"].shift(1) <= df["dea"].shift(1))

    return df


def normalize_daily_cache_df(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()
    if "trade_date" in df.columns:
        df["trade_date"] = df["trade_date"].astype(str)
    return df


def get_daily_cache_file_path(trade_date):
    file_name = f"{str(trade_date)}_{os.path.basename(DAILY_CACHE_CSV)}"
    return os.path.join(os.path.dirname(DAILY_CACHE_CSV), file_name)


def read_daily_cache_file(cache_file):
    if not os.path.exists(cache_file):
        return pd.DataFrame()

    df = pd.read_csv(cache_file, dtype={"ts_code": str, "trade_date": str})
    return normalize_daily_cache_df(df)


def load_daily_cache(trade_dates):
    cached_frames = []
    cached_dates = set()

    for trade_date in trade_dates:
        cache_file = get_daily_cache_file_path(trade_date)
        date_df = read_daily_cache_file(cache_file)

        if date_df.empty:
            continue

        cached_frames.append(date_df)
        cached_dates.add(str(trade_date))

    if not cached_frames:
        return pd.DataFrame(), cached_dates

    cache_df = pd.concat(cached_frames, ignore_index=True)
    cache_df = cache_df.drop_duplicates(subset=["ts_code", "trade_date"], keep="last")
    cache_df = cache_df.sort_values(by=["trade_date", "ts_code"]).reset_index(drop=True)
    return cache_df, cached_dates


def save_daily_cache(df):
    if df is None or df.empty:
        return

    if "trade_date" not in df.columns:
        raise ValueError("daily 缓存缺少 trade_date，无法按交易日拆分保存")

    os.makedirs(os.path.dirname(DAILY_CACHE_CSV), exist_ok=True)
    df = normalize_daily_cache_df(df)
    df = df.drop_duplicates(subset=["ts_code", "trade_date"], keep="last")
    for trade_date, date_df in df.groupby("trade_date"):
        cache_file = get_daily_cache_file_path(trade_date)
        date_df = date_df.sort_values(by=["ts_code"]).reset_index(drop=True)
        date_df.to_csv(cache_file, index=False, encoding="utf-8-sig")


def fetch_daily_by_trade_date(trade_date):
    pro = pro_api()
    return pro.daily(trade_date=trade_date)


def load_moneyflow_cache():
    if not os.path.exists(MONEYFLOW_CACHE_CSV):
        return pd.DataFrame()
    df = pd.read_csv(MONEYFLOW_CACHE_CSV, dtype={"ts_code": str, "trade_date": str})
    if "trade_date" in df.columns:
        df["trade_date"] = df["trade_date"].astype(str)
    return df


def save_moneyflow_cache(df):
    os.makedirs(os.path.dirname(MONEYFLOW_CACHE_CSV), exist_ok=True)
    df = df.drop_duplicates(subset=["ts_code", "trade_date"], keep="last")
    df = df.sort_values(by=["trade_date", "ts_code"]).reset_index(drop=True)
    df.to_csv(MONEYFLOW_CACHE_CSV, index=False, encoding="utf-8-sig")


def fetch_moneyflow_by_trade_date(trade_date):
    pro = pro_api()
    return pro.moneyflow(trade_date=trade_date)


def load_all_moneyflow(trade_dates):
    cache_df = load_moneyflow_cache()
    cached_dates = set(cache_df["trade_date"].astype(str)) if not cache_df.empty and "trade_date" in cache_df.columns else set()

    missing_dates = [trade_date for trade_date in trade_dates if trade_date not in cached_dates]
    if missing_dates:
        print(f"资金流缓存缺失 {len(missing_dates)} 个交易日，开始补齐：{', '.join(missing_dates)}")
    else:
        print(f"资金流缓存已命中最近 {len(trade_dates)} 个交易日，无需重新拉取")

    for trade_date in missing_dates:
        df = fetch_with_retry(
            lambda trade_date=trade_date: fetch_moneyflow_by_trade_date(trade_date),
            f"moneyflow {trade_date}"
        )
        if df.empty:
            raise RuntimeError(f"moneyflow {trade_date} 返回空数据，已停止本次任务，避免使用不完整数据")
        df["trade_date"] = df["trade_date"].astype(str)
        cache_df = pd.concat([cache_df, df], ignore_index=True)
        save_moneyflow_cache(cache_df)
        print(f"已补齐资金流数据：{trade_date}")

    columns = ["ts_code", "trade_date", "external_internal_ratio", "main_net", "main_inflow_2days"]
    if cache_df.empty:
        return pd.DataFrame(columns=columns)

    buy_vol_columns = ["buy_sm_vol", "buy_md_vol", "buy_lg_vol", "buy_elg_vol"]
    sell_vol_columns = ["sell_sm_vol", "sell_md_vol", "sell_lg_vol", "sell_elg_vol"]
    money_amount_columns = ["buy_lg_amount", "buy_elg_amount", "sell_lg_amount", "sell_elg_amount"]

    for col in buy_vol_columns + sell_vol_columns + money_amount_columns:
        if col not in cache_df.columns:
            cache_df[col] = 0
        cache_df[col] = pd.to_numeric(cache_df[col], errors="coerce").fillna(0)

    cache_df["external_vol"] = cache_df[buy_vol_columns].sum(axis=1)
    cache_df["internal_vol"] = cache_df[sell_vol_columns].sum(axis=1)
    cache_df["external_internal_ratio"] = (
        cache_df["external_vol"] / cache_df["internal_vol"].where(cache_df["internal_vol"] > 0)
    )
    cache_df["main_net"] = (
        cache_df["buy_lg_amount"]
        + cache_df["buy_elg_amount"]
        - cache_df["sell_lg_amount"]
        - cache_df["sell_elg_amount"]
    )

    cache_df = cache_df.sort_values(["ts_code", "trade_date"]).reset_index(drop=True)
    cache_df["main_inflow"] = cache_df["main_net"] > 0
    cache_df["prev_main_inflow"] = cache_df.groupby("ts_code")["main_inflow"].shift(1).eq(True)
    cache_df["main_inflow_2days"] = cache_df["main_inflow"] & cache_df["prev_main_inflow"]

    trade_date_set = set(trade_dates)
    return cache_df[cache_df["trade_date"].isin(trade_date_set)][columns].copy()


def load_basic_cache():
    df = eps_downloader.load_basic_cache_file()
    if df.empty:
        return df

    drop_columns = [col for col in ["total_mv", "eps"] if col in df.columns]
    if drop_columns:
        df = df.drop(columns=drop_columns)
    numeric_columns = ["volume_ratio", "turnover_rate", "pe", "dv_ttm"]
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def save_basic_cache(df):
    df = df.drop(columns=["total_mv"], errors="ignore")
    eps_downloader.upsert_basic_cache_rows(df)


def fetch_daily_basic_by_trade_date(trade_date):
    pro = pro_api()
    fields = "ts_code,trade_date,pe,pe_ttm,volume_ratio,turnover_rate,dv_ttm"
    return pro.daily_basic(
        trade_date=trade_date,
        fields=fields
    )


def fetch_total_mv_by_trade_date(trade_date):
    pro = pro_api()
    return pro.daily_basic(
        trade_date=trade_date,
        fields="ts_code,trade_date,total_mv"
    )


def load_latest_total_mv(ts_codes, trade_date):
    if not TOTAL_MV_FILTER_ENABLED:
        return pd.DataFrame(columns=["ts_code", "total_mv"])

    total_mv_df = fetch_with_retry(
        lambda: fetch_total_mv_by_trade_date(trade_date),
        f"daily_basic total_mv {trade_date}"
    )
    if total_mv_df.empty:
        raise RuntimeError(f"daily_basic total_mv {trade_date} 返回空数据，已停止本次任务，避免使用不完整数据")

    total_mv_df = total_mv_df.copy()
    total_mv_df["ts_code"] = total_mv_df["ts_code"].astype(str)
    total_mv_df["total_mv"] = pd.to_numeric(total_mv_df["total_mv"], errors="coerce")
    ts_code_set = set(ts_codes)
    total_mv_df = total_mv_df[total_mv_df["ts_code"].isin(ts_code_set)].copy()
    total_mv_df = total_mv_df.drop_duplicates(subset=["ts_code"], keep="last")
    print(
        f"已拉取最近交易日总市值：{trade_date}，"
        f"{len(total_mv_df)} 只股票"
    )
    return total_mv_df[["ts_code", "total_mv"]]


def merge_basic_frames(daily_basic_df, trade_date):
    daily_basic_columns = ["ts_code", "trade_date", "pe", "pe_ttm", "volume_ratio", "turnover_rate", "dv_ttm"]
    if daily_basic_df is None or daily_basic_df.empty:
        daily_basic_df = pd.DataFrame(columns=daily_basic_columns)

    daily_basic_df = daily_basic_df.copy()

    if "trade_date" not in daily_basic_df.columns:
        daily_basic_df["trade_date"] = trade_date

    merged = daily_basic_df.copy()

    if "pe_ttm" in merged.columns:
        merged["pe"] = merged["pe_ttm"].where(merged["pe_ttm"].notna(), merged.get("pe"))

    columns = ["ts_code", "trade_date", "volume_ratio", "turnover_rate", "pe", "dv_ttm"]
    for col in columns:
        if col not in merged.columns:
            merged[col] = pd.NA

    return merged[columns].copy()


def load_all_basic(ts_codes, trade_dates, daily_df=None):
    cache_df = load_basic_cache()
    required_columns = {"ts_code", "trade_date", "volume_ratio", "turnover_rate", "pe", "dv_ttm"}
    cache_has_required_columns = required_columns.issubset(set(cache_df.columns))
    valid_cached_dates = set()
    invalid_cached_dates = []
    expected_counts = {}
    if daily_df is not None and not daily_df.empty and "trade_date" in daily_df.columns:
        expected_counts = (
            daily_df.assign(trade_date=daily_df["trade_date"].astype(str))
            .groupby("trade_date")["ts_code"]
            .nunique()
            .to_dict()
        )
    if cache_has_required_columns and not cache_df.empty and "trade_date" in cache_df.columns:
        trade_date_cache = cache_df["trade_date"].astype(str)
        for trade_date in trade_dates:
            date_df = cache_df[trade_date_cache == trade_date]
            if date_df.empty:
                continue

            daily_basic_columns = ["volume_ratio", "turnover_rate", "pe", "dv_ttm"]
            daily_basic_ready = date_df[daily_basic_columns].notna().any(axis=0).all()
            expected_count = expected_counts.get(trade_date)
            coverage_ready = True
            if expected_count:
                coverage_ready = date_df["ts_code"].nunique() >= max(1, int(expected_count * 0.9))

            if daily_basic_ready and coverage_ready:
                valid_cached_dates.add(trade_date)
            else:
                invalid_cached_dates.append(trade_date)

    missing_dates = [trade_date for trade_date in trade_dates if trade_date not in valid_cached_dates]
    if missing_dates:
        message = f"基础面缓存缺失或不完整 {len(missing_dates)} 个交易日，开始补齐：{', '.join(missing_dates)}"
        if invalid_cached_dates:
            message += f"；其中缓存不完整日期：{', '.join(invalid_cached_dates)}"
        print(message)
    else:
        print(f"基础面缓存已命中最近 {len(trade_dates)} 个交易日，无需重新拉取")

    for trade_date in missing_dates:
        daily_basic_df = fetch_with_retry(
            lambda trade_date=trade_date: fetch_daily_basic_by_trade_date(trade_date),
            f"daily_basic {trade_date}"
        )
        if daily_basic_df.empty:
            raise RuntimeError(f"daily_basic {trade_date} 返回空数据，已停止本次任务，避免使用不完整数据")
        merged = merge_basic_frames(daily_basic_df, trade_date)
        if not merged.empty:
            cache_df = pd.concat([cache_df, merged], ignore_index=True)
            save_basic_cache(cache_df)
            print(f"已补齐基础面数据：{trade_date}")

    if cache_df.empty:
        columns = ["ts_code", "trade_date", "volume_ratio", "turnover_rate", "pe", "dv_ttm"]
        if EPS_FILTER_ENABLED:
            columns.insert(2, "eps")
        return pd.DataFrame(columns=columns)

    ts_code_set = set(ts_codes)
    trade_date_set = set(trade_dates)
    result = cache_df[
        cache_df["ts_code"].isin(ts_code_set)
        & cache_df["trade_date"].isin(trade_date_set)
    ].copy()
    result_columns = ["ts_code", "trade_date", "volume_ratio", "turnover_rate", "pe"]
    result = result[[col for col in result_columns if col in result.columns]].copy()
    if EPS_FILTER_ENABLED:
        # EPS 改为独立下载；主流程只读取 basic 缓存文件里已准备好的结果，不再顺手触发 bak_basic 下载。
        eps_cache_df = eps_downloader.load_eps_cache()
        if not eps_cache_df.empty:
            eps_result = eps_cache_df[
                eps_cache_df["ts_code"].isin(ts_code_set)
                & eps_cache_df["trade_date"].isin(trade_date_set)
            ][["ts_code", "trade_date", "eps"]].copy()
            result = result.merge(eps_result, on=["ts_code", "trade_date"], how="left")
        else:
            result["eps"] = pd.NA
        result_columns.insert(2, "eps")
        result = result[[col for col in result_columns if col in result.columns]].copy()
    print(
        f"本次使用基础面缓存：{len(result)} 行，"
        f"{result['trade_date'].nunique() if not result.empty else 0} 个交易日，"
        f"{result['ts_code'].nunique() if not result.empty else 0} 只股票"
    )
    return result


def load_shareholder_cache():
    if not os.path.exists(SHAREHOLDER_CACHE_CSV):
        return pd.DataFrame(columns=["ts_code", "holder_flag", "holder_end_date"])

    return pd.read_csv(
        SHAREHOLDER_CACHE_CSV,
        dtype={"ts_code": str, "holder_flag": bool, "holder_end_date": str}
    )


def save_shareholder_cache(df):
    os.makedirs(os.path.dirname(SHAREHOLDER_CACHE_CSV), exist_ok=True)
    df = df.drop_duplicates(subset=["ts_code"], keep="last")
    df = df.sort_values(by=["ts_code"]).reset_index(drop=True)
    df.to_csv(SHAREHOLDER_CACHE_CSV, index=False, encoding="utf-8-sig")


def is_social_security_holder_name(holder_names):
    # 股东名称不是精确匹配，形如“xxx全国社保基金xxx”也应视为社保基金持仓。
    return holder_names.astype(str).str.contains("全国社保基金", regex=False, na=False).any()


def fetch_social_security_holder_flag(ts_code):
    pro = pro_api()
    sources = []
    for api_name in ["top10_floatholders", "top10_holders"]:
        try:
            df = getattr(pro, api_name)(ts_code=ts_code)
            if df is not None and not df.empty:
                df = df.copy()
                if "holder_name" not in df.columns:
                    continue
                if "end_date" in df.columns:
                    df["end_date"] = df["end_date"].astype(str)
                    latest_end_date = df["end_date"].max()
                    latest_df = df[df["end_date"] == latest_end_date]
                else:
                    latest_end_date = ""
                    latest_df = df
                holder_flag = is_social_security_holder_name(latest_df["holder_name"])
                sources.append((bool(holder_flag), latest_end_date))
        except Exception:
            continue

    if not sources:
        return False, ""

    for holder_flag, holder_end_date in sources:
        if holder_flag:
            return True, holder_end_date
    return False, sources[0][1]


def get_social_security_holder_flag(ts_code, shareholder_cache_ref):
    cache_df = shareholder_cache_ref["df"]
    cached = cache_df[cache_df["ts_code"] == ts_code]
    if not cached.empty:
        return bool(cached.iloc[-1]["holder_flag"])

    holder_flag, holder_end_date = fetch_social_security_holder_flag(ts_code)
    new_row = pd.DataFrame([{
        "ts_code": ts_code,
        "holder_flag": holder_flag,
        "holder_end_date": holder_end_date
    }])
    if cache_df.empty:
        shareholder_cache_ref["df"] = new_row
    else:
        shareholder_cache_ref["df"] = pd.concat([cache_df, new_row], ignore_index=True)
    shareholder_cache_ref["dirty"] = True
    return holder_flag


def is_ttm_dividend_yield_in_range(dv_ttm):
    if pd.isna(dv_ttm):
        return False
    if MIN_DV_TTM is not None and dv_ttm < MIN_DV_TTM:
        return False
    if MAX_DV_TTM is not None and dv_ttm > MAX_DV_TTM:
        return False
    return True


def load_all_daily(ts_codes, trade_dates):
    cache_df, cached_dates = load_daily_cache(trade_dates)

    missing_dates = [trade_date for trade_date in trade_dates if trade_date not in cached_dates]
    if missing_dates:
        print(f"日线缓存缺失 {len(missing_dates)} 个交易日，开始补齐：{', '.join(missing_dates)}")
    else:
        print(f"日线缓存已命中最近 {len(trade_dates)} 个交易日，无需重新拉取 daily")

    for trade_date in missing_dates:
        df = fetch_with_retry(
            lambda trade_date=trade_date: fetch_daily_by_trade_date(trade_date),
            f"daily {trade_date}"
        )
        if df.empty:
            raise RuntimeError(f"daily {trade_date} 返回空数据，已停止本次任务，避免使用不完整数据")
        df["trade_date"] = df["trade_date"].astype(str)
        save_daily_cache(df)
        cache_df = pd.concat([cache_df, df], ignore_index=True)
        print(f"已补齐日线数据：{trade_date}")

    if cache_df.empty:
        return pd.DataFrame()

    ts_code_set = set(ts_codes)
    trade_date_set = set(trade_dates)
    result = cache_df[
        cache_df["ts_code"].isin(ts_code_set)
        & cache_df["trade_date"].isin(trade_date_set)
    ].copy()
    print(
        f"本次使用日线缓存：{len(result)} 行，"
        f"{result['trade_date'].nunique() if not result.empty else 0} 个交易日，"
        f"{result['ts_code'].nunique() if not result.empty else 0} 只股票"
    )
    return result


def add_stat(stats, key):
    if stats is not None:
        stats[key] = stats.get(key, 0) + 1


def ensure_numeric_column(df, col):
    if col not in df.columns:
        df[col] = pd.NA
    df[col] = pd.to_numeric(df[col], errors="coerce")


def get_basic_filter_pe_series(df, prefer_pe_ttm=False):
    ensure_numeric_column(df, "pe")
    if prefer_pe_ttm:
        ensure_numeric_column(df, "pe_ttm")
        return df["pe_ttm"].where(df["pe_ttm"] > 0, df["pe"])
    return df["pe"]


def apply_basic_filters(
    df,
    *,
    stock_name_map=None,
    exclude_st=False,
    exclude_bj=False,
    prefer_pe_ttm=False,
    estimate_eps_when_missing=False,
):
    """
    统一基础过滤入口：
    - 股票池阶段：可直接处理 ST / BJ；
    - 信号候选阶段：在 ST / BJ 之外继续处理 EPS / PE / 总市值。
    """
    if df.empty:
        summary = {
            "filter_names": [],
            "before_stock_count": 0,
            "after_stock_count": 0,
            "before_row_count": 0,
            "after_row_count": 0,
        }
        return df.copy(), summary

    filtered_df = df.copy()
    if stock_name_map is not None and "name" not in filtered_df.columns and "ts_code" in filtered_df.columns:
        filtered_df["name"] = filtered_df["ts_code"].map(stock_name_map)

    filter_names = []
    eligible_mask = pd.Series(True, index=filtered_df.index, dtype=bool)

    if exclude_st and "name" in filtered_df.columns:
        eligible_mask &= ~filtered_df["name"].astype(str).str.contains("ST", na=False)
        filter_names.append("ST")

    if exclude_bj and "ts_code" in filtered_df.columns:
        eligible_mask &= ~filtered_df["ts_code"].astype(str).str.endswith(".BJ")
        filter_names.append("BJ")

    has_basic_price_context = "close" in filtered_df.columns
    if has_basic_price_context:
        ensure_numeric_column(filtered_df, "close")

    if EPS_FILTER_ENABLED and has_basic_price_context:
        ensure_numeric_column(filtered_df, "eps")
        if estimate_eps_when_missing:
            pe_ref = get_basic_filter_pe_series(filtered_df, prefer_pe_ttm=prefer_pe_ttm)
            eps_estimated = filtered_df["close"] / pe_ref.where(pe_ref > 0)
            filtered_df["eps"] = filtered_df["eps"].where(filtered_df["eps"].notna(), eps_estimated)
        eligible_mask &= filtered_df["eps"].isna() | (filtered_df["eps"] >= MIN_EPS)
        filter_names.append("EPS")

    if PE_FILTER_ENABLED and "pe" in filtered_df.columns:
        pe_for_filter = get_basic_filter_pe_series(filtered_df, prefer_pe_ttm=prefer_pe_ttm)
        eligible_mask &= pe_for_filter.isna() | ((pe_for_filter > MIN_PE) & (pe_for_filter <= MAX_PE))
        filter_names.append("PE")

    if TOTAL_MV_FILTER_ENABLED and "total_mv" in filtered_df.columns:
        ensure_numeric_column(filtered_df, "total_mv")
        eligible_mask &= filtered_df["total_mv"].isna() | (filtered_df["total_mv"] > MIN_TOTAL_MV)
        filter_names.append("总市值")

    result_df = filtered_df.loc[eligible_mask].copy()
    summary = {
        "filter_names": filter_names,
        "before_stock_count": filtered_df["ts_code"].nunique() if "ts_code" in filtered_df.columns else len(filtered_df),
        "after_stock_count": result_df["ts_code"].nunique() if "ts_code" in result_df.columns else len(result_df),
        "before_row_count": len(filtered_df),
        "after_row_count": len(result_df),
    }
    return result_df, summary


def print_basic_filter_summary(summary, context_label):
    filter_desc = " + ".join(summary["filter_names"]) if summary["filter_names"] else "无"
    print(
        f"{context_label}基础过滤完成（{filter_desc}）："
        f"股票数 {summary['before_stock_count']} -> {summary['after_stock_count']}；"
        f"记录数 {summary['before_row_count']} -> {summary['after_row_count']}"
    )


def build_industry_relative_valuation_result(
    df,
    pe_column="pe",
    min_pe=0.0,
    max_pe=MAX_PE,
    min_sample_count=INDUSTRY_PE_MIN_SAMPLE_COUNT,
    max_percentile=INDUSTRY_PE_MAX_PERCENTILE,
    max_ratio_to_median=INDUSTRY_PE_MAX_RATIO_TO_MEDIAN,
):
    pe_series = pd.to_numeric(df.get(pe_column), errors="coerce")
    stock_count = pd.to_numeric(df.get("industry_stock_count"), errors="coerce")
    percentile = pd.to_numeric(df.get("industry_pe_percentile"), errors="coerce")
    ratio_to_median = pd.to_numeric(df.get("industry_pe_ratio_to_median"), errors="coerce")

    passed = pd.Series(True, index=df.index, dtype=bool)
    reason = pd.Series("通过", index=df.index, dtype=object)

    pe_missing = pe_series.isna()
    reason = reason.mask(pe_missing, "PE缺失，跳过行业估值过滤")

    pe_below_min = pe_series.notna() & (pe_series < min_pe)
    reason = reason.mask(pe_below_min, "市盈率小于0")
    passed &= ~pe_below_min

    pe_above_max = pe_series.notna() & (pe_series > max_pe)
    absolute_pe_invalid = passed & pe_above_max
    reason = reason.mask(absolute_pe_invalid, "市盈率不达标")
    passed &= ~pe_above_max

    pe_needs_industry_eval = passed & pe_series.notna()

    metrics_ready = stock_count.notna() & percentile.notna() & ratio_to_median.notna()
    metrics_missing = pe_needs_industry_eval & ~metrics_ready
    reason = reason.mask(metrics_missing, "行业估值数据不足")
    passed &= ~pe_needs_industry_eval | metrics_ready

    sample_small = stock_count < min_sample_count
    fallback_mask = pe_needs_industry_eval & metrics_ready & sample_small
    reason = reason.mask(fallback_mask, "行业样本不足，回退绝对PE")

    percentile_too_high = pe_needs_industry_eval & metrics_ready & ~sample_small & (percentile > max_percentile)
    reason = reason.mask(percentile_too_high, "行业估值分位不够低")
    passed &= ~pe_needs_industry_eval | sample_small | (percentile <= max_percentile)

    ratio_too_high = pe_needs_industry_eval & metrics_ready & ~sample_small & (ratio_to_median > max_ratio_to_median)
    reason = reason.mask(ratio_too_high, "相对行业中位数折价不足")
    passed &= ~pe_needs_industry_eval | sample_small | (ratio_to_median <= max_ratio_to_median)

    return pd.DataFrame({
        "industry_relative_valuation_low": passed,
        "industry_relative_valuation_reason": reason,
    }, index=df.index)


def add_industry_valuation_metrics(all_daily, stock_pool):
    if all_daily.empty:
        return all_daily

    industry_df = stock_pool[["ts_code", "industry"]].copy()
    merged = all_daily.merge(industry_df, on="ts_code", how="left")
    merged["industry"] = merged["industry"].fillna("").astype(str)
    merged["pe"] = pd.to_numeric(merged["pe"], errors="coerce")

    valid_pe_mask = (
        merged["industry"].ne("")
        & merged["industry"].ne("None")
        & merged["pe"].notna()
        & (merged["pe"] > 0)
    )
    valid_pe_df = merged.loc[valid_pe_mask, ["trade_date", "industry", "ts_code", "pe"]].copy()

    metric_columns = [
        "trade_date",
        "industry",
        "ts_code",
        "industry_stock_count",
        "industry_pe_mean",
        "industry_pe_median",
        "industry_pe_percentile",
        "industry_pe_ratio_to_mean",
        "industry_pe_ratio_to_median",
        "industry_pe_discount_to_mean_pct",
        "industry_pe_discount_to_median_pct",
    ]
    if valid_pe_df.empty:
        for col in metric_columns[3:]:
            merged[col] = pd.NA
        valuation_result = build_industry_relative_valuation_result(merged)
        merged["industry_relative_valuation_low"] = valuation_result["industry_relative_valuation_low"]
        merged["industry_relative_valuation_reason"] = valuation_result["industry_relative_valuation_reason"]
        return merged

    industry_stats = (
        valid_pe_df.groupby(["trade_date", "industry"])["pe"]
        .agg(
            industry_stock_count="count",
            industry_pe_mean="mean",
            industry_pe_median="median",
        )
        .reset_index()
    )

    valid_pe_df = valid_pe_df.sort_values(["trade_date", "industry", "pe", "ts_code"]).reset_index(drop=True)
    valid_pe_df["industry_pe_rank"] = valid_pe_df.groupby(["trade_date", "industry"]).cumcount() + 1
    valid_pe_df = valid_pe_df.merge(industry_stats, on=["trade_date", "industry"], how="left")
    valid_pe_df["industry_pe_percentile"] = valid_pe_df["industry_pe_rank"] / valid_pe_df["industry_stock_count"]
    valid_pe_df["industry_pe_ratio_to_mean"] = valid_pe_df["pe"] / valid_pe_df["industry_pe_mean"].where(
        valid_pe_df["industry_pe_mean"] > 0
    )
    valid_pe_df["industry_pe_ratio_to_median"] = valid_pe_df["pe"] / valid_pe_df["industry_pe_median"].where(
        valid_pe_df["industry_pe_median"] > 0
    )
    valid_pe_df["industry_pe_discount_to_mean_pct"] = (
        1 - valid_pe_df["industry_pe_ratio_to_mean"]
    ) * 100
    valid_pe_df["industry_pe_discount_to_median_pct"] = (
        1 - valid_pe_df["industry_pe_ratio_to_median"]
    ) * 100

    merged = merged.merge(valid_pe_df[metric_columns], on=["trade_date", "industry", "ts_code"], how="left")
    valuation_result = build_industry_relative_valuation_result(merged)
    merged["industry_relative_valuation_low"] = valuation_result["industry_relative_valuation_low"]
    merged["industry_relative_valuation_reason"] = valuation_result["industry_relative_valuation_reason"]
    return merged


def check_signal(
    df,
    i,
    ts_code,
    trade_dates,
    shareholder_cache_ref,
    stats=None
):
    row = df.iloc[i]
    prev = df.iloc[i - 1]

    if CONDITION_FLAGS["trend_above_ma20"] and pd.isna(row["ma20"]):
        add_stat(stats, "指标数据不足")
        return False
    if CONDITION_FLAGS["bullish_ma_alignment"] and (
        pd.isna(row["ma5"]) or pd.isna(row["ma10"]) or pd.isna(row["ma20"])
    ):
        add_stat(stats, "指标数据不足")
        return False
    if CONDITION_FLAGS["volume_rule"] and pd.isna(row["vol_ma5"]):
        add_stat(stats, "指标数据不足")
        return False
    if CONDITION_FLAGS["position_rule"] and (
        pd.isna(row["ma10"]) or pd.isna(row["ma20"]) or pd.isna(row["vol_ma5"])
    ):
        add_stat(stats, "指标数据不足")
        return False
    if CONDITION_FLAGS["rsi_not_overheated"] and pd.isna(row["rsi"]):
        add_stat(stats, "指标数据不足")
        return False
    if CONDITION_FLAGS["volume_ratio_high"] and pd.isna(row["volume_ratio"]):
        add_stat(stats, "基础面数据不足")
        return False
    if CONDITION_FLAGS["turnover_rate_range"] and pd.isna(row["turnover_rate"]):
        add_stat(stats, "基础面数据不足")
        return False
    if CONDITION_FLAGS["prev_year_high_dividend"] and pd.isna(row.get("dv_ttm", pd.NA)):
        add_stat(stats, "基础面数据不足")
        return False
    if CONDITION_FLAGS["external_internal_ratio_high"] and pd.isna(row["external_internal_ratio"]):
        add_stat(stats, "资金流数据不足")
        return False
    if CONDITION_FLAGS["main_money_inflow_2days"] and pd.isna(row["main_inflow_2days"]):
        add_stat(stats, "资金流数据不足")
        return False

    # 1. 股价在20日线之上
    if CONDITION_FLAGS["trend_above_ma20"]:
        cond_trend_1 = row["close"] > row["ma20"]
        if not cond_trend_1:
            add_stat(stats, "股价未站上20日线")
            return False

    # 2. 5日 > 10日 > 20日
    if CONDITION_FLAGS["bullish_ma_alignment"]:
        cond_trend_2 = row["ma5"] > row["ma10"] > row["ma20"]
        if not cond_trend_2:
            add_stat(stats, "均线未多头排列")
            return False

    # 3. 上涨放量 or 回调缩量
    price_up = row["close"] > prev["close"]
    price_down = row["close"] < prev["close"]
    vol_up = row["vol"] > row["vol_ma5"]
    vol_down = row["vol"] < row["vol_ma5"]
    if CONDITION_FLAGS["volume_rule"]:
        cond_volume = (price_up and vol_up) or (price_down and vol_down)
        if not cond_volume:
            add_stat(stats, "量价条件不满足")
            return False

    # 4. 回踩10/20日线缩量，或突破前高/压力位放量
    if CONDITION_FLAGS["position_rule"]:
        near_ma10 = abs(row["close"] - row["ma10"]) / row["ma10"] <= NEAR_MA_THRESHOLD
        near_ma20 = abs(row["close"] - row["ma20"]) / row["ma20"] <= NEAR_MA_THRESHOLD
        pullback_shrink = (near_ma10 or near_ma20) and vol_down
        recent_high = df.iloc[max(0, i - RECENT_HIGH_LOOKBACK_DAYS):i]["high"].max()
        breakout = row["close"] > recent_high and vol_up
        cond_position = pullback_shrink or breakout
        if not cond_position:
            add_stat(stats, "位置条件不满足")
            return False

    # 5. MACD金叉
    if CONDITION_FLAGS["macd_golden_cross"]:
        cond_macd = row["macd_gold"]
        if not cond_macd:
            add_stat(stats, "MACD未金叉")
            return False

    # 6. RSI < 70
    if CONDITION_FLAGS["rsi_not_overheated"]:
        cond_rsi = row["rsi"] < RSI_MAX
        if not cond_rsi:
            add_stat(stats, "RSI过热")
            return False

    # 7. 换手量比达标
    if CONDITION_FLAGS["volume_ratio_high"]:
        if row["volume_ratio"] < MIN_VOLUME_RATIO:
            add_stat(stats, "换手量比不达标")
            return False

    # 8. 外盘 / 内盘达标
    if CONDITION_FLAGS["external_internal_ratio_high"]:
        if row["external_internal_ratio"] < MIN_EXTERNAL_INTERNAL_RATIO:
            add_stat(stats, "外盘内盘比不达标")
            return False

    # 9. 换手率在合理区间
    if CONDITION_FLAGS["turnover_rate_range"]:
        if row["turnover_rate"] < MIN_TURNOVER_RATE or row["turnover_rate"] > MAX_TURNOVER_RATE:
            add_stat(stats, "换手率不达标")
            return False

    # 10. 相对所属行业处于低估值区间
    if CONDITION_FLAGS["industry_relative_valuation_low"]:
        if not bool(row.get("industry_relative_valuation_low", False)):
            add_stat(stats, row.get("industry_relative_valuation_reason", "行业估值不达标"))
            return False
        if row.get("industry_relative_valuation_reason") == "行业样本不足，回退绝对PE":
            add_stat(stats, "行业样本不足，回退绝对PE")

    # 11. 股东成分包含全国社保基金
    if CONDITION_FLAGS["social_security_holder"]:
        if not get_social_security_holder_flag(ts_code, shareholder_cache_ref):
            add_stat(stats, "股东成分不含全国社保基金")
            return False

    # 12. TTM 股息率在目标区间
    if CONDITION_FLAGS["prev_year_high_dividend"]:
        if not is_ttm_dividend_yield_in_range(row.get("dv_ttm", pd.NA)):
            add_stat(stats, "TTM股息率不在目标区间")
            return False

    # 13. 主力资金连续流入2天
    if CONDITION_FLAGS["main_money_inflow_2days"]:
        cond_money = bool(row["main_inflow_2days"])
        if not cond_money:
            add_stat(stats, "主力资金未连续流入")
            return False

    add_stat(stats, "筛选通过")
    return True


# ======================
# 主逻辑
# ======================
def load_stock_pool(ts_codes=None, exclude_st=None, exclude_bj=False):
    stock_pool = basic_api.stock_basic()
    if ts_codes:
        stock_pool = stock_pool[stock_pool["ts_code"].isin(ts_codes)]
    if exclude_st is None:
        exclude_st = RUNTIME_CONFIG["exclude_st_stocks"]
    stock_pool, _ = apply_basic_filters(
        stock_pool,
        exclude_st=exclude_st,
        exclude_bj=exclude_bj,
    )
    return stock_pool.reset_index(drop=True)


def choose_strategy(stock_pool=None, end_date=END_DATE):
    if stock_pool is None:
        stock_pool = load_stock_pool()

    if DATA_LOOKBACK_TRADE_DAYS < SIGNAL_DAYS:
        raise ValueError("DATA_LOOKBACK_TRADE_DAYS 不能小于 SIGNAL_DAYS")

    moneyflow_needed = (
        CONDITION_FLAGS["external_internal_ratio_high"]
        or CONDITION_FLAGS["main_money_inflow_2days"]
    )
    if end_date is None:
        end_date = get_latest_completed_trade_date(require_moneyflow=moneyflow_needed)
    elif moneyflow_needed:
        end_date = resolve_trade_date_with_required_data(end_date, require_moneyflow=True)

    enabled_conditions = [key for key, value in CONDITION_FLAGS.items() if value and key != "pe_reasonable"]
    print(
        f"当前策略：{RUNTIME_CONFIG['strategy_name']}，"
        f"{RUNTIME_CONFIG.get('description', '')}"
    )
    enabled_conditions_text = ", ".join(enabled_conditions) if enabled_conditions else "仅基础过滤"
    print(f"启用条件：{enabled_conditions_text}")
    print(f"统计截止交易日：{end_date}")

    trade_dates = get_trade_dates(end_date, count=DATA_LOOKBACK_TRADE_DAYS)
    if not trade_dates:
        raise ValueError("未获取到可用交易日，请检查 end_date")

    signal_dates = trade_dates[-SIGNAL_DAYS:]
    print(
        f"交易日窗口：{trade_dates[0]} ~ {trade_dates[-1]}，"
        f"共 {len(trade_dates)} 个交易日；筛选日期：{', '.join(signal_dates)}"
    )

    stock_info = {
        row["ts_code"]: row["name"]
        for _, row in stock_pool.iterrows()
    }
    ts_codes = list(stock_info.keys())
    all_daily = load_all_daily(ts_codes, trade_dates)
    if all_daily.empty:
        return pd.DataFrame()
    if TOTAL_MV_FILTER_ENABLED:
        # 总市值过滤按最近一个交易日口径执行，不跟随历史信号日回溯。
        latest_total_mv_df = load_latest_total_mv(ts_codes, trade_dates[-1])
        all_daily = all_daily.merge(latest_total_mv_df, on=["ts_code"], how="left")
    all_basic = load_all_basic(ts_codes, trade_dates, daily_df=all_daily)
    if not all_basic.empty:
        all_daily = all_daily.merge(
            all_basic,
            on=["ts_code", "trade_date"],
            how="left"
        )
    required_basic_columns = ["volume_ratio", "turnover_rate", "pe", "dv_ttm"]
    if EPS_FILTER_ENABLED:
        required_basic_columns.append("eps")
    for col in required_basic_columns:
        if col not in all_daily.columns:
            all_daily[col] = pd.NA
    if TOTAL_MV_FILTER_ENABLED and "total_mv" not in all_daily.columns:
        all_daily["total_mv"] = pd.NA
    all_daily = add_industry_valuation_metrics(all_daily, stock_pool)
    signal_df = all_daily[all_daily["trade_date"].astype(str).isin(set(signal_dates))].copy()
    filtered_signal_df, basic_filter_summary = apply_basic_filters(
        signal_df,
        stock_name_map=stock_info,
        exclude_st=RUNTIME_CONFIG["exclude_st_stocks"],
        prefer_pe_ttm=False,
    )
    print_basic_filter_summary(basic_filter_summary, "筛选信号窗口")
    eligible_signal_keys = set(zip(
        filtered_signal_df["ts_code"].astype(str),
        filtered_signal_df["trade_date"].astype(str),
    ))
    if not eligible_signal_keys:
        return pd.DataFrame()
    if moneyflow_needed:
        # 资金流条件依赖信号日前一日，用完整交易日窗口预计算可避免逐股逐日请求接口。
        all_moneyflow = load_all_moneyflow(trade_dates)
        if not all_moneyflow.empty:
            all_daily = all_daily.merge(
                all_moneyflow,
                on=["ts_code", "trade_date"],
                how="left"
            )
    for col in ["external_internal_ratio", "main_inflow_2days"]:
        if col not in all_daily.columns:
            all_daily[col] = pd.NA

    results = []
    stats = {}
    shareholder_cache_ref = {
        "df": load_shareholder_cache() if CONDITION_FLAGS["social_security_holder"] else pd.DataFrame(),
        "dirty": False
    }
    if CONDITION_FLAGS["prev_year_high_dividend"]:
        print(
            f"TTM股息率筛选区间：{MIN_DV_TTM}% ~ {MAX_DV_TTM}%"
        )

    for ts_code, raw_df in all_daily.groupby("ts_code"):
        try:
            df = add_daily_indicators(raw_df)
            if df.empty or len(df) < 2:
                add_stat(stats, "日线数据不足")
                continue

            name = stock_info.get(ts_code, "")
            for signal_date in signal_dates:
                matched = df[df["trade_date"] == signal_date]
                if matched.empty:
                    add_stat(stats, "筛选日无日线")
                    continue

                if (str(ts_code), str(signal_date)) not in eligible_signal_keys:
                    add_stat(stats, "基础过滤未通过")
                    continue

                i = matched.index[0]
                if i < 1:
                    add_stat(stats, "缺少前一交易日")
                    continue

                signal_row = df.loc[i]

                if not check_signal(
                    df,
                    i,
                    ts_code,
                    trade_dates,
                    shareholder_cache_ref,
                    stats
                ):
                    continue

                base_close = df.loc[i, "close"]
                base_high = df.loc[i, "high"]
                base_low = df.loc[i, "low"]

                results.append({
                    "ts_code": ts_code,
                    "name": name,
                    "industry": signal_row.get("industry", ""),
                    "signal_date": signal_date,
                    "pe": signal_row.get("pe"),
                    "industry_stock_count": signal_row.get("industry_stock_count"),
                    "industry_pe_mean": signal_row.get("industry_pe_mean"),
                    "industry_pe_median": signal_row.get("industry_pe_median"),
                    "industry_pe_percentile": signal_row.get("industry_pe_percentile"),
                    "industry_pe_ratio_to_mean": signal_row.get("industry_pe_ratio_to_mean"),
                    "industry_pe_ratio_to_median": signal_row.get("industry_pe_ratio_to_median"),
                    "industry_pe_discount_to_mean_pct": signal_row.get("industry_pe_discount_to_mean_pct"),
                    "industry_pe_discount_to_median_pct": signal_row.get("industry_pe_discount_to_median_pct"),
                    "base_close": base_close,
                    "base_high": base_high,
                    "base_low": base_low,
                    "forward_day": 0,
                    "future_date": signal_date,

                    "future_close": base_close,
                    "close_diff": 0,
                    "close_pct": 0,

                    "future_high": base_high,
                    "high_diff": 0,
                    "high_pct": 0,

                    "future_low": base_low,
                    "low_diff": 0,
                    "low_pct": 0,
                })

                available_forward_days = len(df) - i - 1
                forward_days = available_forward_days
                if MAX_FORWARD_DAYS is not None:
                    forward_days = min(available_forward_days, MAX_FORWARD_DAYS)

                for n in range(1, forward_days + 1):
                    if i + n >= len(df):
                        continue

                    future = df.loc[i + n]

                    results.append({
                        "ts_code": ts_code,
                        "name": name,
                        "industry": signal_row.get("industry", ""),
                        "signal_date": signal_date,
                        "pe": signal_row.get("pe"),
                        "industry_stock_count": signal_row.get("industry_stock_count"),
                        "industry_pe_mean": signal_row.get("industry_pe_mean"),
                        "industry_pe_median": signal_row.get("industry_pe_median"),
                        "industry_pe_percentile": signal_row.get("industry_pe_percentile"),
                        "industry_pe_ratio_to_mean": signal_row.get("industry_pe_ratio_to_mean"),
                        "industry_pe_ratio_to_median": signal_row.get("industry_pe_ratio_to_median"),
                        "industry_pe_discount_to_mean_pct": signal_row.get("industry_pe_discount_to_mean_pct"),
                        "industry_pe_discount_to_median_pct": signal_row.get("industry_pe_discount_to_median_pct"),
                        "base_close": base_close,
                        "base_high": base_high,
                        "base_low": base_low,
                        "forward_day": n,
                        "future_date": future["trade_date"],

                        "future_close": future["close"],
                        "close_diff": future["close"] - base_close,
                        "close_pct": (future["close"] / base_close - 1) * 100,

                        "future_high": future["high"],
                        "high_diff": future["high"] - base_high,
                        "high_pct": (future["high"] / base_high - 1) * 100,

                        "future_low": future["low"],
                        "low_diff": future["low"] - base_low,
                        "low_pct": (future["low"] / base_low - 1) * 100,
                    })

        except Exception as e:
            print(f"{ts_code} error: {e}")

    if shareholder_cache_ref["dirty"]:
        save_shareholder_cache(shareholder_cache_ref["df"])

    print("筛选诊断：")
    for key, value in stats.items():
        print(f"- {key}: {value}")

    return pd.DataFrame(results)


def format_price(value):
    return f"{value:.2f}"


def format_diff(value):
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.2f}"


def init_display_row(columns):
    return {column: "" for column in columns}


def get_enabled_condition_names():
    return [
        CONDITION_NAMES.get(key, key)
        for key, enabled in CONDITION_FLAGS.items()
        if enabled and key != "pe_reasonable"
    ]


def build_display_table(detail_df):
    base_df = detail_df[detail_df["forward_day"] == 0].copy()
    future_df = detail_df[detail_df["forward_day"] > 0].copy()

    signal_dates = set(base_df["signal_date"])
    dates = sorted(set(detail_df["signal_date"]).union(set(detail_df["future_date"])))
    fixed_columns = ["股票代码", "股票名称", "信号日期"]
    columns = fixed_columns.copy()
    for date in dates:
        columns.append(date)
        if date in signal_dates:
            columns.append(f"{date}_价格")

    rows = []
    block_index = {}

    base_df = base_df.sort_values(by=["signal_date", "ts_code"]).reset_index(drop=True)
    for _, row in base_df.iterrows():
        start_row = len(rows)
        block_index[(row["signal_date"], row["ts_code"])] = start_row

        for _ in range(3):
            rows.append(init_display_row(columns))

        for offset in range(3):
            rows[start_row + offset]["股票代码"] = row["ts_code"]
            rows[start_row + offset]["股票名称"] = row["name"]
            rows[start_row + offset]["信号日期"] = row["signal_date"]

        date_col = row["signal_date"]
        price_col = f"{row['signal_date']}_价格"

        rows[start_row][price_col] = f"收盘价{format_price(row['base_close'])}"
        rows[start_row + 1][price_col] = f"最高价{format_price(row['base_high'])}"
        rows[start_row + 2][price_col] = f"最低价{format_price(row['base_low'])}"

    future_df = future_df.sort_values(
        by=["signal_date", "ts_code", "forward_day"]
    ).reset_index(drop=True)

    for _, row in future_df.iterrows():
        start_row = block_index.get((row["signal_date"], row["ts_code"]))
        if start_row is None:
            continue

        date_col = row["future_date"]
        rows[start_row][date_col] = format_diff(row["close_diff"])
        rows[start_row + 1][date_col] = format_diff(row["high_diff"])
        rows[start_row + 2][date_col] = format_diff(row["low_diff"])

    return pd.DataFrame(rows, columns=columns)


def build_valuation_compare_table(detail_df):
    compare_columns = [
        "signal_date",
        "ts_code",
        "name",
        "industry",
        "pe",
        "industry_stock_count",
        "industry_pe_mean",
        "industry_pe_median",
        "industry_pe_percentile",
        "industry_pe_ratio_to_mean",
        "industry_pe_ratio_to_median",
        "industry_pe_discount_to_mean_pct",
        "industry_pe_discount_to_median_pct",
    ]
    if detail_df.empty or not set(compare_columns).issubset(detail_df.columns):
        return pd.DataFrame()

    base_df = detail_df[detail_df["forward_day"] == 0][compare_columns].copy()
    if base_df.empty:
        return pd.DataFrame()

    base_df = base_df.sort_values(
        by=["signal_date", "industry_pe_percentile", "industry_pe_ratio_to_median", "pe", "ts_code"],
        ascending=[False, True, True, True, True]
    ).reset_index(drop=True)
    base_df["industry_pe_percentile"] = base_df["industry_pe_percentile"] * 100
    base_df["industry_pe_ratio_to_mean"] = base_df["industry_pe_ratio_to_mean"] * 100
    base_df["industry_pe_ratio_to_median"] = base_df["industry_pe_ratio_to_median"] * 100

    rename_map = {
        "signal_date": "信号日期",
        "ts_code": "股票代码",
        "name": "股票名称",
        "industry": "行业板块",
        "pe": "个股PE",
        "industry_stock_count": "行业样本数",
        "industry_pe_mean": "行业平均PE",
        "industry_pe_median": "行业中位PE",
        "industry_pe_percentile": "行业PE分位(%)",
        "industry_pe_ratio_to_mean": "个股/行业平均PE(%)",
        "industry_pe_ratio_to_median": "个股/行业中位PE(%)",
        "industry_pe_discount_to_mean_pct": "较行业平均折价(%)",
        "industry_pe_discount_to_median_pct": "较行业中位折价(%)",
    }
    compare_df = base_df.rename(columns=rename_map)
    numeric_columns = [col for col in compare_df.columns if col not in ["信号日期", "股票代码", "股票名称", "行业板块"]]
    compare_df[numeric_columns] = compare_df[numeric_columns].apply(pd.to_numeric, errors="coerce").round(2)
    return compare_df


def fill_dataframe_sheet(worksheet, dataframe):
    if dataframe is None or dataframe.empty:
        worksheet["A1"] = "无数据"
        return

    for col_index, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=column_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[get_column_letter(col_index)].width = max(12, len(str(column_name)) + 4)

    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"


def save_display_excel(display_table, valuation_compare_table=None, filename=RESULT_XLSX):
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "策略筛选"
    summary_sheet = workbook.create_sheet("策略说明")
    valuation_sheet = workbook.create_sheet("估值对比")
    red_font = Font(color="FF0000")
    green_font = Font(color="008000")

    enabled_condition_names = get_enabled_condition_names()
    summary_sheet["A1"] = "当前策略"
    summary_sheet["B1"] = RUNTIME_CONFIG["strategy_name"]
    summary_sheet["A2"] = "策略说明"
    summary_sheet["B2"] = RUNTIME_CONFIG.get("description", "")
    summary_sheet["A3"] = "基础过滤"
    eps_filter_text = f"EPS >= {MIN_EPS}" if EPS_FILTER_ENABLED else "EPS不过滤"
    pe_filter_text = f"PE 在 (0, {MAX_PE}] 且空值放行" if PE_FILTER_ENABLED else "PE不过滤"
    total_mv_filter_text = f"总市值 > {MIN_TOTAL_MV / 10000:.0f}亿" if TOTAL_MV_FILTER_ENABLED else "总市值不过滤"
    summary_sheet["B3"] = f"排除ST：{RUNTIME_CONFIG['exclude_st_stocks']}；{eps_filter_text}；{pe_filter_text}；{total_mv_filter_text}"
    summary_sheet["A4"] = "启用条件"
    summary_sheet["B4"] = "、".join(enabled_condition_names) if enabled_condition_names else "仅基础过滤"
    summary_sheet["A5"] = "行业估值阈值"
    summary_sheet["B5"] = (
        f"行业样本数 >= {INDUSTRY_PE_MIN_SAMPLE_COUNT}；"
        f"行业PE分位 <= {INDUSTRY_PE_MAX_PERCENTILE * 100:.0f}%；"
        f"个股/行业中位PE <= {INDUSTRY_PE_MAX_RATIO_TO_MEDIAN * 100:.0f}%"
    )
    summary_sheet["A6"] = "启用条件明细"
    for row_index, condition_name in enumerate(enabled_condition_names, start=7):
        summary_sheet.cell(row=row_index, column=1, value=condition_name)
    for col_index in range(1, 3):
        summary_sheet.column_dimensions[get_column_letter(col_index)].width = 36

    column_count = len(display_table.columns)
    col_index = 1
    while col_index <= column_count:
        date = display_table.columns[col_index - 1]
        has_price_column = (
            col_index < column_count
            and display_table.columns[col_index] == f"{date}_价格"
        )
        if has_price_column:
            worksheet.merge_cells(
                start_row=1,
                start_column=col_index,
                end_row=1,
                end_column=col_index + 1
            )
        cell = worksheet.cell(row=1, column=col_index, value=date)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_index += 2 if has_price_column else 1

    for row_index, row in enumerate(display_table.itertuples(index=False), start=2):
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="center")
            if isinstance(value, str) and value.startswith("+"):
                cell.font = red_font
            elif isinstance(value, str) and value.startswith("-"):
                cell.font = green_font

    worksheet.freeze_panes = "D2"
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 12
    for col_index in range(4, column_count + 1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = 18

    fill_dataframe_sheet(valuation_sheet, valuation_compare_table)
    workbook.save(filename)
    return True


def save_result_tables(detail_df):
    if detail_df.empty:
        print(f"最近 {SIGNAL_DAYS} 个交易日没有筛选出符合条件的股票")
        return None

    display_table = build_display_table(detail_df)
    valuation_compare_table = build_valuation_compare_table(detail_df)
    saved_xlsx = save_display_excel(display_table, valuation_compare_table=valuation_compare_table)

    if saved_xlsx:
        print(f"策略筛选横向展示表：{RESULT_XLSX}")
    print("\n===== 策略筛选横向展示表 =====")
    print(display_table)

    return display_table


def main():
    detail_df = choose_strategy()
    save_result_tables(detail_df)


if __name__ == '__main__':
    main()
