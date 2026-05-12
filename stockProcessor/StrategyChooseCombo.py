import os
from copy import deepcopy

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import StrategyChoose as choose_module
import strategy_choose_config
from stockProcessor.download.constants import score_result_path

# ======================
# 参数区
# ======================
# 这里直接配置两个“条件 key 列表”，风格与 StrategyScore 的 CONDITION_KEYS 一致。
# 每个组合只启用列表里的条件，其余条件关闭；基础过滤仍沿用 choose 默认配置。
STRATEGY_COMBINATIONS = [
    {
        "label": "策略1",
        "condition_keys": [
            "bullish_ma_alignment",
            "position_rule",
            "rsi_not_overheated",
            "volume_ratio_high",
            "turnover_rate_range",
            "industry_relative_valuation_low",
            "social_security_holder",
            "main_money_inflow_2days"
        ],
    },
    {
        "label": "策略2",
        "condition_keys": [
            "bullish_ma_alignment",
            "position_rule",
            "volume_rule",
            "macd_golden_cross",
            "rsi_not_overheated",
            "turnover_rate_range",
            "main_money_inflow_2days"
        ],
# trend_above_ma20: 股价在20日线之上
    # bullish_ma_alignment: 5日 > 10日 > 20日
    # volume_rule: 上涨放量或回调缩量
    # position_rule: 回踩10/20日线缩量，或突破前高放量
    # macd_golden_cross: MACD金叉
    # rsi_not_overheated: RSI 不超过阈值
    # volume_ratio_high: 量比不低于
    # external_internal_ratio_high: 外盘 / 内盘不低于1.05
    # turnover_rate_range: 换手率在设定区间内 1-12
    # pe_reasonable: 启用市盈率基础过滤，要求 PE 在合理区间 0-80；pe 为空时放行
    # industry_relative_valuation_low: 相对所属行业处于低估值区间
    # social_security_holder: 股东成分包含全国社保基金
    # prev_year_high_dividend: 上一年度现金分红较高，10股税前大于1
    # main_money_inflow_2days: 主力资金连续流入2天
    },
]
END_DATE = None  # None 表示自动取最近一个 daily + moneyflow 都已就绪的交易日
TOP_N = 5
RESULT_XLSX = score_result_path("strategy_choose_combo_result.xlsx")


def normalize_optional_number(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return float(value)


def validate_condition_keys(condition_keys):
    valid_keys = set(strategy_choose_config.DEFAULT_CONFIG["conditions"].keys())
    invalid_keys = [key for key in condition_keys if key not in valid_keys]
    if invalid_keys:
        raise ValueError(f"存在无效条件 key: {', '.join(invalid_keys)}")


def build_runtime_config(combo):
    config = deepcopy(strategy_choose_config.DEFAULT_CONFIG)
    condition_keys = combo.get("condition_keys", [])
    validate_condition_keys(condition_keys)
    config["conditions"] = {
        key: key in condition_keys
        for key in config["conditions"].keys()
    }
    config["strategy_name"] = combo["label"]
    config["description"] = "条件组合：" + "、".join(
        choose_module.CONDITION_NAMES.get(key, key)
        for key in condition_keys
    ) if condition_keys else "仅基础过滤"
    return config


def apply_runtime_config(config):
    """
    StrategyChoose 以模块级全局变量驱动。
    这里统一覆盖相关全局值，使新脚本可以无侵入地复用原有下载和筛选逻辑。
    """
    choose_module.RUNTIME_CONFIG = config
    choose_module.MARKET_CLOSE_HOUR = int(config.get("market_close_hour", 15))
    choose_module.DATA_LOOKBACK_TRADE_DAYS = config["data_lookback_trade_days"]
    choose_module.SIGNAL_DAYS = config["signal_days"]
    choose_module.MAX_FORWARD_DAYS = config["max_forward_days"]
    choose_module.VOL_MA_DAYS = config["vol_ma_days"]
    choose_module.RSI_PERIOD = config["rsi_period"]
    choose_module.PREV_YEAR_MIN_CASH_DIV_TAX = config["prev_year_min_cash_div_tax"]
    choose_module.CONDITION_FLAGS = config["conditions"]
    choose_module.RSI_MAX = config["rsi_max"]
    choose_module.NEAR_MA_THRESHOLD = config["near_ma_threshold"]
    choose_module.RECENT_HIGH_LOOKBACK_DAYS = config["recent_high_lookback_days"]

    choose_module.MIN_EPS = normalize_optional_number(config["min_eps"])
    choose_module.EPS_FILTER_ENABLED = bool(config.get("enable_eps_filter", True)) and choose_module.MIN_EPS is not None
    choose_module.MIN_TOTAL_MV = normalize_optional_number(config.get("min_total_mv"))
    choose_module.TOTAL_MV_FILTER_ENABLED = bool(config.get("enable_total_mv_filter", True)) and choose_module.MIN_TOTAL_MV is not None
    choose_module.MIN_VOLUME_RATIO = config["min_volume_ratio"]
    choose_module.MIN_EXTERNAL_INTERNAL_RATIO = config["min_external_internal_ratio"]
    choose_module.MIN_TURNOVER_RATE = config["min_turnover_rate"]
    choose_module.MAX_TURNOVER_RATE = config["max_turnover_rate"]
    choose_module.MAX_PE = normalize_optional_number(config["max_pe"])
    choose_module.PE_FILTER_ENABLED = bool(config["conditions"].get("pe_reasonable", False)) and choose_module.MAX_PE is not None
    choose_module.INDUSTRY_PE_MIN_SAMPLE_COUNT = int(config.get("industry_pe_min_sample_count", 5))
    choose_module.INDUSTRY_PE_MAX_PERCENTILE = float(config.get("industry_pe_max_percentile", 0.35))
    choose_module.INDUSTRY_PE_MAX_RATIO_TO_MEDIAN = float(config.get("industry_pe_max_ratio_to_median", 0.8))


def resolve_common_end_date(combos):
    first_config = build_runtime_config(combos[0])
    apply_runtime_config(first_config)

    if END_DATE is None:
        return choose_module.get_latest_completed_trade_date(require_moneyflow=True)
    return choose_module.resolve_trade_date_with_required_data(END_DATE, require_moneyflow=True)


def run_strategy_combo(combo, end_date):
    config = build_runtime_config(combo)
    apply_runtime_config(config)

    print(f"\n===== {combo['label']} 开始筛选 =====")
    detail_df = choose_module.choose_strategy(end_date=end_date)
    if detail_df.empty:
        print(f"{combo['label']} 未筛选出结果")
        return detail_df, config

    detail_df = detail_df.copy()
    detail_df["strategy_label"] = combo["label"]
    detail_df["strategy_keys"] = "、".join(combo.get("condition_keys", []))
    detail_df["strategy_description"] = config.get("description", "")
    return detail_df, config


def build_strategy_base_table(detail_df):
    if detail_df.empty:
        return pd.DataFrame()

    base_df = detail_df[detail_df["forward_day"] == 0].copy()
    columns = [
        "signal_date",
        "ts_code",
        "name",
        "industry",
        "strategy_label",
        "strategy_keys",
        "base_close",
        "base_high",
        "base_low",
        "pe",
    ]
    existing_columns = [col for col in columns if col in base_df.columns]
    base_df = base_df[existing_columns].copy()
    rename_map = {
        "signal_date": "信号日期",
        "ts_code": "股票代码",
        "name": "股票名称",
        "industry": "行业",
        "strategy_label": "策略标签",
        "strategy_keys": "条件Key列表",
        "base_close": "信号日收盘价",
        "base_high": "信号日最高价",
        "base_low": "信号日最低价",
        "pe": "PE",
    }
    base_df = base_df.rename(columns=rename_map)
    return base_df.sort_values(by=["信号日期", "股票代码"]).reset_index(drop=True)


def merge_strategy_details(detail_frames, combos):
    non_empty_frames = [df for df in detail_frames if not df.empty]
    if not non_empty_frames:
        return pd.DataFrame(), pd.DataFrame()

    combined_df = pd.concat(non_empty_frames, ignore_index=True)
    base_df = combined_df[combined_df["forward_day"] == 0].copy()

    combo_label_order = {combo["label"]: index for index, combo in enumerate(combos)}
    combo_keys_order = {
        "、".join(combo.get("condition_keys", [])): index
        for index, combo in enumerate(combos)
    }

    def sort_by_order(values, order_map):
        ordered_values = sorted(set(values), key=lambda item: order_map.get(item, 999))
        return "、".join(ordered_values)

    summary_df = (
        base_df.groupby(["signal_date", "ts_code"], as_index=False)
        .agg(
            name=("name", "first"),
            industry=("industry", "first"),
            pe=("pe", "first"),
            source_strategies=("strategy_label", lambda values: sort_by_order(values, combo_label_order)),
            strategy_keys=("strategy_keys", lambda values: sort_by_order(values, combo_keys_order)),
            strategy_hit_count=("strategy_label", "nunique"),
        )
    )

    # 合并口径是并集：
    # 只要某只股票在某个 signal_date 命中任一组合，就保留；
    # 如果两个组合都命中，则按 signal_date + ts_code 去重后保留一条，并记录命中来源。
    dedup_df = combined_df.sort_values(
        by=["signal_date", "ts_code", "forward_day", "future_date", "strategy_label"]
    ).drop_duplicates(
        subset=["signal_date", "ts_code", "forward_day", "future_date"],
        keep="first",
    )

    merged_detail_df = dedup_df.drop(
        columns=["strategy_label", "strategy_keys", "strategy_description"],
        errors="ignore",
    ).merge(
        summary_df,
        on=["signal_date", "ts_code"],
        how="left",
        suffixes=("", "_summary"),
    )

    for column in ["name", "industry", "pe"]:
        summary_column = f"{column}_summary"
        if summary_column in merged_detail_df.columns:
            merged_detail_df[column] = merged_detail_df[column].fillna(merged_detail_df[summary_column])
            merged_detail_df = merged_detail_df.drop(columns=[summary_column])

    merged_base_df = merged_detail_df[merged_detail_df["forward_day"] == 0].copy()
    merged_base_df = merged_base_df.sort_values(
        by=["signal_date", "ts_code"]
    ).reset_index(drop=True)
    return merged_detail_df, merged_base_df


def zscore(series):
    if series is None or len(series) == 0:
        return 0.0

    numeric_series = pd.to_numeric(series, errors="coerce")
    latest_value = numeric_series.iloc[-1]
    if pd.isna(latest_value):
        return 0.0

    mean = numeric_series.mean()
    std = numeric_series.std()
    if std == 0 or np.isnan(std):
        return 0.0
    return float((latest_value - mean) / std)


def clip_score(value):
    return float(np.clip(value, -3, 3))


def calc_stock_score(df):
    close = pd.to_numeric(df["close"], errors="coerce")
    volume = pd.to_numeric(df["volume"], errors="coerce")
    money_series = pd.to_numeric(df["money_flow"], errors="coerce")

    vol_ratio_series = volume / volume.rolling(20).mean()
    volume_score = clip_score(zscore(vol_ratio_series))

    breakout_series = close / close.rolling(20).max()
    breakout_score = clip_score(zscore(breakout_series))

    money_score = clip_score(zscore(money_series))

    ma20 = close.rolling(20).mean()
    ma60 = close.rolling(60).mean()
    trend_series = ma20 / ma60
    trend_score = clip_score(zscore(trend_series))

    score = (
        0.2 * volume_score
        + 0.3 * breakout_score
        + 0.3 * money_score
        + 0.2 * trend_score
    )

    return {
        "volume_score": round(volume_score, 4),
        "breakout_score": round(breakout_score, 4),
        "money_score": round(money_score, 4),
        "trend_score": round(trend_score, 4),
        "score": round(float(score), 4),
    }


def load_score_history(ts_codes, end_date, lookback_trade_days):
    default_config = build_runtime_config(STRATEGY_COMBINATIONS[0])
    apply_runtime_config(default_config)

    trade_dates = choose_module.get_trade_dates(end_date, count=lookback_trade_days)
    daily_df = choose_module.load_all_daily(ts_codes, trade_dates)
    moneyflow_df = choose_module.load_all_moneyflow(trade_dates)

    if daily_df.empty:
        return pd.DataFrame()

    history_df = daily_df[["ts_code", "trade_date", "close", "vol"]].copy()
    history_df = history_df.rename(columns={"vol": "volume"})
    if moneyflow_df.empty:
        history_df["money_flow"] = 0.0
    else:
        history_df = history_df.merge(
            moneyflow_df[["ts_code", "trade_date", "main_net"]],
            on=["ts_code", "trade_date"],
            how="left",
        )
        history_df["money_flow"] = pd.to_numeric(history_df["main_net"], errors="coerce").fillna(0.0)
        history_df = history_df.drop(columns=["main_net"])

    history_df["trade_date"] = history_df["trade_date"].astype(str)
    history_df = history_df.sort_values(by=["ts_code", "trade_date"]).reset_index(drop=True)
    return history_df


def score_merged_stocks(merged_base_df, runtime_configs, end_date):
    if merged_base_df.empty:
        return merged_base_df

    lookback_trade_days = max(
        max(config["data_lookback_trade_days"], 80)
        for config in runtime_configs
    )
    ts_codes = sorted(merged_base_df["ts_code"].astype(str).unique().tolist())
    history_df = load_score_history(ts_codes, end_date, lookback_trade_days)
    if history_df.empty:
        raise RuntimeError("评分所需历史数据为空，无法继续生成合并结果")

    history_map = {
        ts_code: df.reset_index(drop=True)
        for ts_code, df in history_df.groupby("ts_code")
    }

    score_rows = []
    for row in merged_base_df.itertuples(index=False):
        stock_history = history_map.get(str(row.ts_code))
        if stock_history is None:
            score_result = {
                "volume_score": 0.0,
                "breakout_score": 0.0,
                "money_score": 0.0,
                "trend_score": 0.0,
                "score": 0.0,
            }
        else:
            signal_history = stock_history[stock_history["trade_date"] <= str(row.signal_date)].copy()
            score_result = calc_stock_score(signal_history)

        score_rows.append({
            "signal_date": row.signal_date,
            "ts_code": row.ts_code,
            **score_result,
        })

    score_df = pd.DataFrame(score_rows)
    ranked_df = merged_base_df.merge(score_df, on=["signal_date", "ts_code"], how="left")
    ranked_df = ranked_df.sort_values(
        by=["signal_date", "score", "strategy_hit_count", "ts_code"],
        ascending=[True, False, False, True],
    ).reset_index(drop=True)
    ranked_df["rank"] = ranked_df.groupby("signal_date").cumcount() + 1
    ranked_df["is_top_n"] = ranked_df["rank"] <= TOP_N
    return ranked_df


def build_rank_detail_table(ranked_base_df):
    if ranked_base_df.empty:
        return pd.DataFrame()

    columns = [
        "signal_date",
        "rank",
        "ts_code",
        "name",
        "industry",
        "score",
        "volume_score",
        "breakout_score",
        "money_score",
        "trend_score",
        "strategy_hit_count",
        "source_strategies",
        "strategy_keys",
        "pe",
    ]
    detail_df = ranked_base_df[[col for col in columns if col in ranked_base_df.columns]].copy()
    rename_map = {
        "signal_date": "信号日期",
        "rank": "排名",
        "ts_code": "股票代码",
        "name": "股票名称",
        "industry": "行业",
        "score": "总分",
        "volume_score": "放量分",
        "breakout_score": "突破分",
        "money_score": "资金分",
        "trend_score": "趋势分",
        "strategy_hit_count": "命中策略数",
        "source_strategies": "命中策略",
        "strategy_keys": "命中条件Key列表",
        "pe": "PE",
    }
    detail_df = detail_df.rename(columns=rename_map)
    detail_df["总分"] = pd.to_numeric(detail_df["总分"], errors="coerce").round(4)
    for column in ["放量分", "突破分", "资金分", "趋势分", "PE"]:
        if column in detail_df.columns:
            detail_df[column] = pd.to_numeric(detail_df[column], errors="coerce").round(4)
    return detail_df


def build_merged_base_table(ranked_base_df):
    if ranked_base_df.empty:
        return pd.DataFrame()

    columns = [
        "signal_date",
        "ts_code",
        "name",
        "industry",
        "score",
        "rank",
        "strategy_hit_count",
        "source_strategies",
        "strategy_keys",
        "pe",
        "base_close",
        "base_high",
        "base_low",
    ]
    detail_df = ranked_base_df[[col for col in columns if col in ranked_base_df.columns]].copy()
    rename_map = {
        "signal_date": "信号日期",
        "ts_code": "股票代码",
        "name": "股票名称",
        "industry": "行业",
        "score": "总分",
        "rank": "排名",
        "strategy_hit_count": "命中策略数",
        "source_strategies": "命中策略",
        "strategy_keys": "命中条件Key列表",
        "pe": "PE",
        "base_close": "信号日收盘价",
        "base_high": "信号日最高价",
        "base_low": "信号日最低价",
    }
    detail_df = detail_df.rename(columns=rename_map)
    return detail_df.sort_values(by=["信号日期", "排名", "股票代码"]).reset_index(drop=True)


def init_display_row(columns):
    return {column: "" for column in columns}


def format_price(value):
    return f"{value:.2f}"


def format_diff(value):
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.2f}"


def build_topn_display_table(detail_df):
    if detail_df.empty:
        return pd.DataFrame()

    base_df = detail_df[detail_df["forward_day"] == 0].copy()
    future_df = detail_df[detail_df["forward_day"] > 0].copy()

    signal_dates = set(base_df["signal_date"])
    dates = sorted(set(detail_df["signal_date"]).union(set(detail_df["future_date"])))
    fixed_columns = ["信号日期", "排名", "股票代码", "股票名称", "来源策略", "得分"]
    columns = fixed_columns.copy()
    for date in dates:
        columns.append(date)
        if date in signal_dates:
            columns.append(f"{date}_价格")

    rows = []
    block_index = {}

    base_df = base_df.sort_values(by=["signal_date", "rank", "ts_code"]).reset_index(drop=True)
    for _, row in base_df.iterrows():
        start_row = len(rows)
        block_index[(row["signal_date"], row["ts_code"])] = start_row

        for _ in range(3):
            rows.append(init_display_row(columns))

        for offset in range(3):
            rows[start_row + offset]["信号日期"] = row["signal_date"]
            rows[start_row + offset]["排名"] = int(row["rank"])
            rows[start_row + offset]["股票代码"] = row["ts_code"]
            rows[start_row + offset]["股票名称"] = row["name"]
            rows[start_row + offset]["来源策略"] = row["source_strategies"]
            rows[start_row + offset]["得分"] = round(float(row["score"]), 4)

        price_col = f"{row['signal_date']}_价格"
        rows[start_row][price_col] = f"收盘价{format_price(row['base_close'])}"
        rows[start_row + 1][price_col] = f"最高价{format_price(row['base_high'])}"
        rows[start_row + 2][price_col] = f"最低价{format_price(row['base_low'])}"

    future_df = future_df.sort_values(by=["signal_date", "rank", "ts_code", "forward_day"]).reset_index(drop=True)
    for _, row in future_df.iterrows():
        start_row = block_index.get((row["signal_date"], row["ts_code"]))
        if start_row is None:
            continue
        date_col = row["future_date"]
        rows[start_row][date_col] = format_diff(row["close_diff"])
        rows[start_row + 1][date_col] = format_diff(row["high_diff"])
        rows[start_row + 2][date_col] = format_diff(row["low_diff"])

    return pd.DataFrame(rows, columns=columns)


def fill_topn_sheet(worksheet, dataframe, fixed_column_count=6):
    if dataframe is None or dataframe.empty:
        worksheet["A1"] = "无数据"
        return

    red_font = Font(color="FF0000")
    green_font = Font(color="008000")

    for col_index in range(1, fixed_column_count + 1):
        header = dataframe.columns[col_index - 1]
        cell = worksheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    column_count = len(dataframe.columns)
    col_index = fixed_column_count + 1
    while col_index <= column_count:
        date = dataframe.columns[col_index - 1]
        has_price_column = (
            col_index < column_count
            and dataframe.columns[col_index] == f"{date}_价格"
        )
        if has_price_column:
            worksheet.merge_cells(
                start_row=1,
                start_column=col_index,
                end_row=1,
                end_column=col_index + 1,
            )
        cell = worksheet.cell(row=1, column=col_index, value=date)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_index += 2 if has_price_column else 1

    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="center")
            if isinstance(value, str) and value.startswith("+"):
                cell.font = red_font
            elif isinstance(value, str) and value.startswith("-"):
                cell.font = green_font

    worksheet.freeze_panes = "G2"
    column_widths = {
        "A": 12,
        "B": 8,
        "C": 14,
        "D": 16,
        "E": 22,
        "F": 12,
    }
    for column_name, width in column_widths.items():
        worksheet.column_dimensions[column_name].width = width

    for col_index in range(fixed_column_count + 1, column_count + 1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = 18


def build_summary_rows(combos, runtime_configs, end_date, merged_base_df, ranked_base_df):
    rows = [
        {"项目": "合并截止交易日", "值": end_date},
        {"项目": "策略组合数量", "值": len(combos)},
        {"项目": "合并后唯一信号数", "值": len(merged_base_df)},
        {"项目": f"每日Top{TOP_N}总记录数", "值": int(ranked_base_df["is_top_n"].sum()) if not ranked_base_df.empty else 0},
        {"项目": "评分规则", "值": "0.2*放量 + 0.3*突破 + 0.3*资金 + 0.2*趋势，子项统一裁剪到[-3,3]"},
    ]

    for combo, config in zip(combos, runtime_configs):
        rows.append({
            "项目": f"{combo['label']} 条件Key列表",
            "值": "、".join(combo.get("condition_keys", [])),
        })
        rows.append({
            "项目": f"{combo['label']} 说明",
            "值": config.get("description", ""),
        })
        enabled_conditions = [
            choose_module.CONDITION_NAMES.get(key, key)
            for key, enabled in config["conditions"].items()
            if enabled
        ]
        rows.append({
            "项目": f"{combo['label']} 启用条件",
            "值": "、".join(enabled_conditions) if enabled_conditions else "仅基础过滤",
        })

    return pd.DataFrame(rows)


def save_result_excel(
    topn_display_df,
    ranked_base_df,
    merged_base_df,
    strategy_result_frames,
    runtime_configs,
    combos,
    end_date,
):
    os.makedirs(os.path.dirname(RESULT_XLSX), exist_ok=True)

    workbook = Workbook()
    topn_sheet = workbook.active
    topn_sheet.title = f"Top{TOP_N}横向收益"
    rank_sheet = workbook.create_sheet(f"每日Top{TOP_N}")
    merged_sheet = workbook.create_sheet("合并明细")
    summary_sheet = workbook.create_sheet("策略说明")
    strategy_sheets = [
        workbook.create_sheet(f"{combo['label']}入选")
        for combo in combos
    ]

    fill_topn_sheet(topn_sheet, topn_display_df)
    choose_module.fill_dataframe_sheet(rank_sheet, build_rank_detail_table(ranked_base_df[ranked_base_df["is_top_n"]].copy()))
    choose_module.fill_dataframe_sheet(merged_sheet, build_merged_base_table(merged_base_df))
    choose_module.fill_dataframe_sheet(summary_sheet, build_summary_rows(combos, runtime_configs, end_date, merged_base_df, ranked_base_df))

    for sheet, frame in zip(strategy_sheets, strategy_result_frames):
        choose_module.fill_dataframe_sheet(sheet, build_strategy_base_table(frame))

    workbook.save(RESULT_XLSX)


def main():
    if len(STRATEGY_COMBINATIONS) != 2:
        raise ValueError("STRATEGY_COMBINATIONS 必须且只能配置两个策略组合")

    end_date = resolve_common_end_date(STRATEGY_COMBINATIONS)
    strategy_frames = []
    runtime_configs = []

    for combo in STRATEGY_COMBINATIONS:
        detail_df, config = run_strategy_combo(combo, end_date)
        strategy_frames.append(detail_df)
        runtime_configs.append(config)

    merged_detail_df, merged_base_df = merge_strategy_details(strategy_frames, STRATEGY_COMBINATIONS)
    if merged_base_df.empty:
        print("两个策略组合都没有筛选出结果，未生成横向收益表")
        return

    ranked_base_df = score_merged_stocks(merged_base_df, runtime_configs, end_date)
    topn_base_df = ranked_base_df[ranked_base_df["is_top_n"]].copy()

    topn_key_df = topn_base_df[["signal_date", "ts_code"]].drop_duplicates().copy()
    topn_detail_df = merged_detail_df.merge(
        topn_key_df,
        on=["signal_date", "ts_code"],
        how="inner",
    )
    topn_detail_df = topn_detail_df.merge(
        topn_base_df[[
            "signal_date",
            "ts_code",
            "rank",
            "score",
            "source_strategies",
        ]],
        on=["signal_date", "ts_code", "source_strategies"],
        how="left",
    )

    topn_display_df = build_topn_display_table(topn_detail_df)
    save_result_excel(
        topn_display_df=topn_display_df,
        ranked_base_df=ranked_base_df,
        merged_base_df=ranked_base_df,
        strategy_result_frames=strategy_frames,
        runtime_configs=runtime_configs,
        combos=STRATEGY_COMBINATIONS,
        end_date=end_date,
    )

    print(f"\n合并策略结果已输出：{RESULT_XLSX}")
    print(f"合并后唯一信号数：{len(merged_base_df)}")
    print(f"每日 Top{TOP_N} 记录数：{len(topn_base_df)}")


if __name__ == "__main__":
    main()
